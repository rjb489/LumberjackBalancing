import pandas as pd
import numpy as np
import openpyxl
import math
import time
import random
from PyQt6.QtCore import QThread, pyqtSignal
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import PatternFill
from collections import defaultdict, Counter
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import PatternFill
from collections import defaultdict

from algorithmPolicy import (
    loadWorkloadPolicy, loadInstructorTrack, loadSpecialCourses,
    rowIsValid, Course, FacultyMember, adjust_co_convened
)

class ExcelProcessor(QThread):
    """Threaded Excel workload processor using updated algorithm."""
    progress = pyqtSignal(int)
    completed = pyqtSignal(str)
    error = pyqtSignal(str)

    def __init__(self, raw_file_path, policy_file_path, track_file_path, special_file_path):
        super().__init__()
        self.raw_file_path = raw_file_path
        self.policy_file_path = policy_file_path
        self.track_file_path = track_file_path
        self.special_file_path = special_file_path



    def run(self):
        try:
            # 1) Load raw data
            converters = {
                'Max Units': lambda x: float(x) if pd.notna(x) else 0.0,
                'Enroll Total': lambda x: int(float(x)) if pd.notna(x) else 0
            }
            raw_df = pd.read_excel(self.raw_file_path, sheet_name='Raw Data', converters=converters)
            raw_df = raw_df[raw_df.apply(rowIsValid, axis=1)].reset_index(drop=True)
            raw_df = raw_df.drop_duplicates(subset=[
                'Instructor Emplid', 'Term', 'Subject', 'Cat Nbr', 'Section',
                'Start Date', 'End Date', 'Start Time', 'End Time', 'Facility Building', 'Facility Room', 'Days'
            ])
            
            # 2) Supporting data
            policy = loadWorkloadPolicy(self.policy_file_path) if self.policy_file_path else loadWorkloadPolicy()
            tracks = loadInstructorTrack(self.track_file_path) if self.track_file_path else {}
            special = loadSpecialCourses(self.special_file_path) if self.special_file_path else set()

            # 3) Build structures
            faculty = {}
            courseGroups = {}
            other = {}

            for _, row in raw_df.iterrows():
                role = str(row.get('Instructor Role', '')).strip().upper()

                emplid_val = row.get('Instructor Emplid')
                if pd.isna(emplid_val):
                    continue

                emplid = int(float(emplid_val))
                if emplid not in tracks:
                    other.setdefault(str(emplid), []).append(row.to_dict())
                    continue

                course = Course(row.to_dict(), policy, special)
                key = course.getGroupKeyForGrouping()
                courseGroups.setdefault(key, []).append(course)

                if emplid not in faculty:
                    faculty[emplid] = FacultyMember(row.get('Instructor', ''), row.get('Instructor Email', ''), emplid, role, tracks[emplid])
                faculty[emplid].addCourse(course)

            # 4) Team‑taught division
            for lst in courseGroups.values():
                valid = [c for c in lst if all(c._meeting_signature())]
                pi_only = [c for c in valid if c.instructorRole.upper()=="PI"]
                unique_emplids = {c.instructorEmplid for c in pi_only}
                if len(unique_emplids) >= 2:
                    names = [c.rawData.get('Instructor','').strip() for c in pi_only]
                    for c in pi_only:
                        c.team_taught_members = [
                            n for n in names 
                            if n != c.rawData.get('Instructor','').strip()
                        ]
                        c.adjustLoadDivision(len(unique_emplids))

            # 5) Co‑convened adjustment
            adjust_co_convened([c for lst in courseGroups.values() for c in lst])

            # 6) Calculate summary
            summary_rows = []
            for fac in faculty.values():
                fac.calculateTotalLoad()
                units = sorted({getattr(c, 'unit', '') for c in fac.courses.values() if getattr(c, 'unit', '')})

                course_list = []
                for c in fac.courses.values():
                    # base label
                    subject = c.rawData.get('Subject','').strip()
                    section = c.rawData.get('Section','').strip()
                    desc    = c.rawData.get('Class Description','').strip().title()
                    label   = f"{subject} {c.catNbr}-{section} – {desc}"

                    # tag on any partner info
                    if getattr(c, 'co_convened_members', None):
                        label += f" (co‑convened with {', '.join(c.co_convened_members)})"
                    if getattr(c, 'team_taught_members', None):
                        label += f" (team‑taught with {', '.join(c.team_taught_members)})"

                    course_list.append(label)

                course_list.sort()
                summary_rows.append({
                    'Instructor': fac.name,
                    'Emplid': fac.emplid,
                    'Track': fac.track or 'Unknown',
                    'Total Workload': round(fac.totalLoad, 2),
                    'Units Taught': ', '.join(units),
                    'Courses Taught': '; '.join(course_list)
                })

            summary_df = pd.DataFrame(summary_rows)

            # 7) Write output
            out_file = self.raw_file_path.replace('.xlsx', '_summary.xlsx')
            with pd.ExcelWriter(out_file, engine='openpyxl') as writer:
                raw_df.to_excel(writer, sheet_name='Processed Raw Data', index=False)
                summary_df.to_excel(writer, sheet_name='Faculty Summary', index=False)
            
            export_faculty_by_unit(faculty, outputFile="faculty_by_unit.xlsx")

            # Simulate progress
            pct = 0
            while pct < 100:
                pct = min(pct + random.randint(5, 10), 100)
                self.progress.emit(pct)
                time.sleep(0.03)

            self.completed.emit(out_file)


        except Exception as e:
            self.error.emit(str(e))

def export_faculty_by_unit(facultyDict, outputFile="faculty_by_unit.xlsx"):

    # Define colors
    GREEN  = "90EE90"   # CT or CT Well
    YELLOW = "FFFF00"   # For intermediary ranges (used in cell fill)
    RED    = "FF6347"   # TT below/bad or CT very low
    ORANGE = "FFA500"   # TT high (or TT poor performance)
    BLUE   = "1E90FF"   # TT or TT Well

    # Define PatternFills
    green_fill  = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
    yellow_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
    red_fill    = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
    orange_fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")

    def table_cell_fill(track_str, displayed_load_val):
        """
        Table cell color logic:
         - If ceiled load is within ±2 of baseline -> green.
         - Else, for loads deviating by 3-6 -> yellow.
         - For loads 7 or more above (or 7 or more below) baseline,
           use orange for high values (CT) and red for low values.
        """
        track_str = (track_str or "").strip().upper()
        if track_str == "CT":
            expected = 40
        elif track_str == "TT":
            expected = 30
        else:
            return red_fill

        diff = displayed_load_val - expected
        if abs(diff) <= 2:
            return green_fill
        elif diff > 0:
            if 3 <= diff <= 6:
                return yellow_fill
            elif diff >= 7:
                return orange_fill
        else:  # diff < 0
            if 3 <= abs(diff) <= 6:
                return yellow_fill
            elif abs(diff) >= 7:
                return red_fill
        return red_fill

    # Chart 1: Baseline Pie Chart (CT=40 vs TT=30)
    def add_simple_pie_chart(ws, anchor_cell="K2", chart_title="CT=40 vs TT=30 (Baseline)"):
        ws["Z2"] = "CT Expectation"
        ws["AA2"] = 40
        ws["Z3"] = "TT Expectation"
        ws["AA3"] = 30

        chart = PieChart()
        chart.title = chart_title
        data_ref = Reference(ws, min_col=27, min_row=2, max_row=3)  
        cat_ref = Reference(ws, min_col=26, min_row=2, max_row=3)   
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(cat_ref)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showCatName = False
        chart.dataLabels.showVal = False
        chart.series[0].data_points = [
            DataPoint(idx=0, spPr=GraphicalProperties(solidFill=GREEN)),
            DataPoint(idx=1, spPr=GraphicalProperties(solidFill=BLUE))
        ]
        ws.add_chart(chart, anchor_cell)

    # Chart 2: Performance Breakdown Pie Chart
    def add_breakdown_pie_chart(ws, anchor_cell, ct_well, ct_other, tt_well, tt_other, chart_title):
        total = ct_well + ct_other + tt_well + tt_other
        if total == 0:
            pct_ct_well = pct_ct_other = pct_tt_well = pct_tt_other = 0
        else:
            pct_ct_well  = round(100 * ct_well / total, 2)
            pct_ct_other = round(100 * ct_other / total, 2)
            pct_tt_well  = round(100 * tt_well / total, 2)
            pct_tt_other = round(100 * tt_other / total, 2)
        ws["AC10"] = "Category"
        ws["AE10"] = "Percentage"
        categories = ["CT Well", "CT Other", "TT Well", "TT Other"]
        values = [pct_ct_well, pct_ct_other, pct_tt_well, pct_tt_other]
        row_ptr = 11
        for label, val in zip(categories, values):
            ws.cell(row=row_ptr, column=29, value=label)  # Column AC.
            ws.cell(row=row_ptr, column=31, value=val)     # Column AE.
            row_ptr += 1
        pie = PieChart()
        pie.title = chart_title
        data_r = Reference(ws, min_col=31, min_row=11, max_row=14)
        cat_r  = Reference(ws, min_col=29, min_row=11, max_row=14)
        pie.add_data(data_r, titles_from_data=False)
        pie.set_categories(cat_r)
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showVal = True
        pie.series[0].data_points = [
            DataPoint(idx=0, spPr=GraphicalProperties(solidFill=GREEN)),
            DataPoint(idx=1, spPr=GraphicalProperties(solidFill=RED)),
            DataPoint(idx=2, spPr=GraphicalProperties(solidFill=BLUE)),
            DataPoint(idx=3, spPr=GraphicalProperties(solidFill=ORANGE))
        ]
        ws.add_chart(pie, anchor_cell)

    # Create workbook and remove default sheet.
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    # ---------------------------
    # 1) Create the "ALL" Sheet with Two Sub-tables (CT in A-C, TT in F-H)
    # ---------------------------
    all_ws = wb.create_sheet("ALL")
    # CT Table header
    all_ws["A1"] = "CT Table"
    all_ws["A2"] = "Name"
    all_ws["B2"] = "Track"
    all_ws["C2"] = "Load"
    # TT Table header
    all_ws["F1"] = "TT Table"
    all_ws["F2"] = "Name"
    all_ws["G2"] = "Track"
    all_ws["H2"] = "Load"
    # Add Baseline Chart (Chart 1)
    add_simple_pie_chart(all_ws, anchor_cell="K2", chart_title="CT=40 vs TT=30 (Baseline)")
    
    # Initialize breakdown counts for ALL sheet.
    ct_well = ct_other = tt_well = tt_other = 0
    row_ct = 3  # For CT table
    row_tt = 3  # For TT table
    all_fac_list = sorted(facultyDict.values(), key=lambda f: f.totalLoad if f.totalLoad else 0.0, reverse=True)
    for fac in all_fac_list:
        load_val = fac.totalLoad
        track_str = (fac.track or "").strip().upper()
        displayed_val = int(math.ceil(load_val))
        if track_str == "CT":
            all_ws.cell(row=row_ct, column=1, value=fac.name)
            all_ws.cell(row=row_ct, column=2, value=fac.track)
            cell_load = all_ws.cell(row=row_ct, column=3, value=displayed_val)
            cell_load.fill = table_cell_fill(track_str, displayed_val)
            if abs(displayed_val - 40) <= 2:
                ct_well += 1
            else:
                ct_other += 1
            row_ct += 1
        elif track_str == "TT":
            all_ws.cell(row=row_tt, column=6, value=fac.name)
            all_ws.cell(row=row_tt, column=7, value=fac.track)
            cell_load = all_ws.cell(row=row_tt, column=8, value=displayed_val)
            cell_load.fill = table_cell_fill(track_str, displayed_val)
            if abs(displayed_val - 30) <= 2:
                tt_well += 1
            else:
                tt_other += 1
            row_tt += 1
    # Add Breakdown Chart (Chart 2)
    add_breakdown_pie_chart(all_ws, anchor_cell="K15", ct_well=ct_well, ct_other=ct_other, tt_well=tt_well, tt_other=tt_other,
                            chart_title="Performance Breakdown (Within ±2 vs Others)")

    # ---------------------------
    # 2) Create a Sheet per Unit (same structure as "ALL")
    # ---------------------------
    unit_map = defaultdict(set)
    for fac in facultyDict.values():
        units = { c.unit.strip() for c in fac.courses.values() if c.unit.strip() }
        for unit in units:
            unit_map[unit].add(fac)
    for unit_name, fac_set in unit_map.items():
        ws = wb.create_sheet(unit_name[:31])
        # CT Table header
        ws["A1"] = "CT Table"
        ws["A2"] = "Name"
        ws["B2"] = "Track"
        ws["C2"] = "Load"
        # TT Table header
        ws["F1"] = "TT Table"
        ws["F2"] = "Name"
        ws["G2"] = "Track"
        ws["H2"] = "Load"
        # Add Baseline Chart for Unit
        add_simple_pie_chart(ws, anchor_cell="K2", chart_title=f"{unit_name}: CT=40 vs TT=30 (Baseline)")
        # Initialize unit breakdown counts.
        ct_well_u = ct_other_u = tt_well_u = tt_other_u = 0
        sorted_facs = sorted(fac_set, key=lambda f: f.totalLoad if f.totalLoad else 0.0, reverse=True)
        row_ct = 3
        row_tt = 3
        for f2 in sorted_facs:
            load_val = f2.totalLoad
            track_s = (f2.track or "").strip().upper()
            displayed_val = int(math.ceil(load_val))
            if track_s == "CT":
                ws.cell(row=row_ct, column=1, value=f2.name)
                ws.cell(row=row_ct, column=2, value=f2.track)
                cell_load = ws.cell(row=row_ct, column=3, value=displayed_val)
                cell_load.fill = table_cell_fill(track_s, displayed_val)
                if abs(displayed_val - 40) <= 2:
                    ct_well_u += 1
                else:
                    ct_other_u += 1
                row_ct += 1
            elif track_s == "TT":
                ws.cell(row=row_tt, column=6, value=f2.name)
                ws.cell(row=row_tt, column=7, value=f2.track)
                cell_load = ws.cell(row=row_tt, column=8, value=displayed_val)
                cell_load.fill = table_cell_fill(track_s, displayed_val)
                if abs(displayed_val - 30) <= 2:
                    tt_well_u += 1
                else:
                    tt_other_u += 1
                row_tt += 1
        add_breakdown_pie_chart(ws, anchor_cell="K15", ct_well=ct_well_u, ct_other=ct_other_u,
                                tt_well=tt_well_u, tt_other=tt_other_u,
                                chart_title=f"{unit_name}: Performance Breakdown")
    wb.save(outputFile)
    print(f"Export complete. See '{outputFile}'.")