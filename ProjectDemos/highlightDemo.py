# Demo to show proof of highlighting analyzed data
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

# Define thresholds
GREEN_THRESHOLD = 85
YELLOW_THRESHOLD = 70

# Read Excel file
input_file = r"C:\Users\riley\OneDrive\Desktop\NAU Stuff\NAU Semester 5\SE476\Coding\LumberjackBalancing\highlightDemoFile1.xlsx"
df = pd.read_excel(input_file)

# Ensure the correct column name
target_column = "Workload Percentage"  # Update this to the column name you're analyzing
if target_column not in df.columns:
    raise KeyError(f"Column '{target_column}' not found in the Excel file. Please check your column names.")

# Analyze data and determine colors
def get_color(value):
    if value >= GREEN_THRESHOLD:
        return "green"
    elif YELLOW_THRESHOLD <= value < GREEN_THRESHOLD:
        return "yellow"
    else:
        return "red"

# Create a dictionary to store colors without adding a new column to the DataFrame
colors = df[target_column].apply(get_color)

# Create a new workbook for the final output
wb = Workbook()
ws = wb.active

# Write headers
for col_idx, col_name in enumerate(df.columns, start=1):
    ws.cell(row=1, column=col_idx, value=col_name)

# Write data and apply formatting
fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fill_yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
fill_red = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")

for row_idx, row in df.iterrows():
    for col_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)
        if col_idx == df.columns.get_loc(target_column) + 1:  # Check if the column is the target column
            color = colors.iloc[row_idx]
            if color == "green":
                cell.fill = fill_green
            elif color == "yellow":
                cell.fill = fill_yellow
            elif color == "red":
                cell.fill = fill_red

# Save final file
output_file = "output.xlsx"
wb.save(output_file)

print(f"Analysis complete. Output saved to {output_file}")