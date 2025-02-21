import tkinter as tk
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk
import os
import sys
from openpyxl import load_workbook, Workbook


def select_file():
    """Open a file dialog to select an Excel file."""
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    if not file_path:
        return  # User canceled file selection

    try:
        process_excel(file_path)
        messagebox.showinfo("Success", "Calculations applied. Output file created!")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to process the file: {e}")


#Open and extracts policy information from an excel file
def calculation_file():
    """Open a file dialog to select an Excel file."""
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    if not file_path:
        return  # User canceled file selection

    #logic to extract policies from excel


def process_excel(input_file):
    """Process the Excel file, apply calculations, and save output."""
    # Load the input workbook
    workbook = load_workbook(input_file)
    sheet = workbook.active

    # Extract data (this could vary depending on the file structure)
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    # Example transformation: Add a sum of the first two columns (optional logic)
    processed_data = []
    for row in data:
        if len(row) >= 2 and all(isinstance(cell, (int, float)) for cell in row[:2]):
            processed_data.append(row + (row[0] + row[1],))
        else:
            processed_data.append(row)

    # Create a new workbook for the output
    output_workbook = Workbook()
    output_sheet = output_workbook.active

    # Write processed data into the output workbook
    for row in processed_data:
        output_sheet.append(row)

    # Save the new Excel file
    output_file = os.path.join(os.path.dirname(input_file), "output.xlsx")
    output_workbook.save(output_file)


def get_image_path(filename):
    """Get the path to the image, handling both script and PyInstaller executables."""
    if hasattr(sys, "_MEIPASS"):
        # PyInstaller temporary directory
        return os.path.join(sys._MEIPASS, filename)
    else:
        # Regular development environment
        return os.path.join(os.path.dirname(__file__), filename)


# Tkinter GUI
root = tk.Tk()
root.title("Excel Processor")
root.geometry("500x400")  
root.configure(bg="yellow")

frame = tk.Frame(root, padx=100, pady=100)
frame.pack()
frame.configure(bg="RoyalBlue")

# Load and display the image
try:
    image_path = get_image_path("Logo.png")
    if not os.path.exists(image_path):
        raise FileNotFoundError(f"Image file '{image_path}' not found.")

    # Load the image and resize it
    image = Image.open(image_path)
    image = image.resize((125, 150), Image.ANTIALIAS)
    photo = ImageTk.PhotoImage(image)

    # Display the image in the GUI
    image_label = tk.Label(frame, image=photo)
    image_label.photo = photo  # Prevent garbage collection
    image_label.pack(pady=10)
except Exception as e:
    print(f"Error loading image: {e}")

label = tk.Label(frame,
                 text="Select an Excel file to process:",
                 fg="black", bg="white")
label.pack(pady=10)

# White buttons
browse_button = tk.Button(frame, text="Browse", bg="white", fg="black",
                          command=select_file)
browse_button.pack()

policy_button = tk.Button(frame, text="Policy Table", bg="white", fg="black",
                          command=calculation_file)
policy_button.pack(pady=10)

exit_button = tk.Button(frame, text="Exit", bg="white", fg="black",
                        command=root.quit)
exit_button.pack()

root.mainloop()
